#! python3
#! spreadsheetRepeatsChecker.py - checks a spreadsheet for repeated values
#   usage - the program will ask for a filename. It will check that spreadsheet for repeated values
#			and print those values for the user to delete if s/he wishes.
import openpyxl
from openpyxl import load_workbook
file_name = input ('What is the name of the file you would like to check? ')

wb = openpyxl.load_workbook(file_name)

sheet = wb['sheet']

x = sheet.cell(row = 1, column = 1).value


values = []

#checks all rows
for i in range (sheet.max_row-1):
	#checks all columns in all rows
	for j in range (sheet.max_column-1):
		cell_value =  (sheet.cell(row = i+1, column = j+1).value)

		#prevents the program from printing a bunch of 'none's
		if cell_value is not None:
			
			#prints the repeated values so the user can go back and delete them if s/he wishes
			if cell_value in values:
				print (cell_value)
			#appends value to list that is checked for repeats
			else:
				values.append(cell_value)
input('Press Enter to exit the program')


